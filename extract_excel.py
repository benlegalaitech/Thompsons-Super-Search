#!/usr/bin/env python3
"""
Excel Text Extraction Script

Extracts text from Excel files (.xls, .xlsx, .xlsm, .xlsb) sheet-by-sheet
and saves as JSON for fast searching. Produces the same index format as
extract.py (for PDFs), so both file types are searchable through the same engine.

Usage:
    python extract_excel.py                    # Run extraction using config.json
    python extract_excel.py --source /path     # Override source folder
    python extract_excel.py --reindex          # Force re-extract all files
"""

import os
import sys
import json
import argparse
from datetime import datetime
from pathlib import Path

from tqdm import tqdm


EXCEL_EXTENSIONS = ['.xls', '.xlsx', '.xlsm', '.xlsb']
MAX_ROWS_PER_SHEET = 500


def load_config():
    """Load configuration from config.json."""
    config_path = Path('config.json')
    if not config_path.exists():
        print("Error: config.json not found.")
        print("Create config.json from config.example.json and set your excel_source_folder.")
        sys.exit(1)

    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def find_excel_files(source_folder, extensions=None):
    """Find all Excel files in the source folder recursively."""
    if extensions is None:
        extensions = EXCEL_EXTENSIONS

    files = []
    source_path = Path(source_folder)

    if not source_path.exists():
        print(f"Error: Source folder does not exist: {source_folder}")
        sys.exit(1)

    for ext in extensions:
        files.extend(source_path.rglob(f'*{ext}'))

    return sorted(files)


def detect_headers(first_row):
    """Check if a row looks like a header row (mostly non-empty strings, not numbers)."""
    if not first_row:
        return None

    non_empty = [c for c in first_row if c is not None and str(c).strip()]
    if not non_empty:
        return None

    string_count = sum(1 for c in non_empty if isinstance(c, str))
    # Consider it a header if at least 60% of non-empty cells are strings
    if string_count / len(non_empty) >= 0.6:
        return [str(c).strip() if c is not None else '' for c in first_row]

    return None


def serialize_row(row, headers=None):
    """Convert a row to searchable text.

    With headers: 'Header1: value1 | Header2: value2'
    Without headers: 'value1 | value2 | value3'
    """
    if headers:
        pairs = []
        for i, cell in enumerate(row):
            if cell is None or str(cell).strip() == '':
                continue
            header = headers[i] if i < len(headers) and headers[i] else ''
            value = str(cell).strip()
            if header:
                pairs.append(f"{header}: {value}")
            else:
                pairs.append(value)
        return ' | '.join(pairs)
    else:
        values = [str(c).strip() for c in row if c is not None and str(c).strip()]
        return ' | '.join(values)


def extract_sheet_openpyxl(ws, max_rows=MAX_ROWS_PER_SHEET):
    """Extract text from an openpyxl worksheet."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(tuple(row))
        # Read a bit beyond max_rows to know if truncation is needed
        if len(rows) > max_rows + 1:
            break

    if not rows:
        return ''

    # Detect headers
    headers = detect_headers(rows[0])
    if headers:
        data_rows = rows[1:max_rows + 1]
        total_rows = len(rows) - 1  # minus header
    else:
        data_rows = rows[:max_rows]
        total_rows = len(rows)

    lines = []
    for row in data_rows:
        line = serialize_row(row, headers)
        if line:
            lines.append(line)

    text = '\n'.join(lines)

    if total_rows > max_rows:
        text += f'\n[... additional rows truncated, showing first {max_rows} of {total_rows}+ rows]'

    return text


def extract_sheet_xlrd(ws, max_rows=MAX_ROWS_PER_SHEET):
    """Extract text from an xlrd worksheet."""
    if ws.nrows == 0:
        return ''

    # Read all rows (xlrd loads them all in memory anyway)
    rows = []
    for r in range(ws.nrows):
        row = tuple(ws.cell_value(r, c) for c in range(ws.ncols))
        rows.append(row)

    # Detect headers
    headers = detect_headers(rows[0])
    if headers:
        data_rows = rows[1:max_rows + 1]
        total_rows = ws.nrows - 1
    else:
        data_rows = rows[:max_rows]
        total_rows = ws.nrows

    lines = []
    for row in data_rows:
        line = serialize_row(row, headers)
        if line:
            lines.append(line)

    text = '\n'.join(lines)

    if total_rows > max_rows:
        text += f'\n[... additional rows truncated, showing first {max_rows} of {total_rows} rows]'

    return text


def extract_sheet_pyxlsb(ws, max_rows=MAX_ROWS_PER_SHEET):
    """Extract text from a pyxlsb worksheet."""
    rows = []
    for row in ws.rows():
        values = tuple(c.v for c in row)
        rows.append(values)
        if len(rows) > max_rows + 1:
            break

    if not rows:
        return ''

    headers = detect_headers(rows[0])
    if headers:
        data_rows = rows[1:max_rows + 1]
        total_rows = len(rows) - 1
    else:
        data_rows = rows[:max_rows]
        total_rows = len(rows)

    lines = []
    for row in data_rows:
        line = serialize_row(row, headers)
        if line:
            lines.append(line)

    text = '\n'.join(lines)

    if total_rows > max_rows:
        text += f'\n[... additional rows truncated, showing first {max_rows} of {total_rows}+ rows]'

    return text


def extract_text_from_excel(filepath):
    """Extract text from all sheets of an Excel file.

    Returns list of {'page_num': N, 'sheet_name': name, 'text': text}
    or None on error.
    """
    ext = filepath.suffix.lower()
    pages = []

    try:
        if ext in ('.xlsx', '.xlsm'):
            import openpyxl
            from openpyxl.chartsheet import Chartsheet
            wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
            try:
                for i, sheet_name in enumerate(wb.sheetnames, start=1):
                    ws = wb[sheet_name]
                    # Skip chart-only sheets (no tabular data)
                    if isinstance(ws, Chartsheet):
                        continue
                    text = extract_sheet_openpyxl(ws)
                    pages.append({
                        'page_num': i,
                        'sheet_name': sheet_name,
                        'text': text.strip()
                    })
            finally:
                wb.close()

        elif ext == '.xls':
            import xlrd
            wb = xlrd.open_workbook(str(filepath))
            for i, sheet_name in enumerate(wb.sheet_names(), start=1):
                ws = wb.sheet_by_name(sheet_name)
                text = extract_sheet_xlrd(ws)
                pages.append({
                    'page_num': i,
                    'sheet_name': sheet_name,
                    'text': text.strip()
                })

        elif ext == '.xlsb':
            from pyxlsb import open_workbook
            with open_workbook(str(filepath)) as wb:
                for i, sheet_name in enumerate(wb.sheets, start=1):
                    with wb.get_sheet(sheet_name) as ws:
                        text = extract_sheet_pyxlsb(ws)
                        pages.append({
                            'page_num': i,
                            'sheet_name': sheet_name,
                            'text': text.strip()
                        })

        else:
            print(f"\nUnsupported extension: {ext} for {filepath}")
            return None

    except Exception as e:
        print(f"\nError extracting {filepath}: {e}")
        return None

    return pages


def get_relative_path(file_path, source_folder):
    """Get the relative path of a file from the source folder."""
    try:
        return str(Path(file_path).relative_to(source_folder))
    except ValueError:
        return str(file_path)


def extract_all(source_folder, index_folder, reindex=False, extensions=None):
    """Extract text from all Excel files and save to index folder."""
    source_path = Path(source_folder)
    index_path = Path(index_folder)
    texts_path = index_path / 'texts'

    # Create folders
    texts_path.mkdir(parents=True, exist_ok=True)

    # Find all Excel files
    print(f"Scanning for Excel files in: {source_folder}")
    files = find_excel_files(source_folder, extensions)
    print(f"Found {len(files)} Excel files")

    if not files:
        print("No Excel files found.")
        return

    # Track statistics
    total_sheets = 0
    processed = 0
    skipped = 0
    errors = 0

    # Process each file
    for filepath in tqdm(files, desc="Extracting text", unit="file"):
        filename = filepath.name
        # Use {stem}.{ext}.json to avoid collisions with PDF index files
        output_file = texts_path / f"{filepath.name}.json"

        # Skip if already extracted (unless reindex)
        if output_file.exists() and not reindex:
            skipped += 1
            continue

        # Extract text
        pages = extract_text_from_excel(filepath)

        if pages is None:
            errors += 1
            continue

        # Save to JSON
        doc = {
            'filename': filename,
            'path': get_relative_path(filepath, source_folder),
            'file_type': 'excel',
            'pages': pages
        }

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(doc, f, ensure_ascii=False, indent=2)

        total_sheets += len(pages)
        processed += 1

    # Save metadata (append to existing if present)
    metadata_file = index_path / 'metadata.json'
    metadata = {}
    if metadata_file.exists():
        with open(metadata_file, 'r', encoding='utf-8') as f:
            metadata = json.load(f)

    metadata['excel_source_folder'] = str(source_folder)
    metadata['excel_total_docs'] = processed + skipped
    metadata['excel_total_sheets'] = total_sheets
    metadata['excel_extracted_at'] = datetime.now().isoformat()
    metadata['excel_processed'] = processed
    metadata['excel_skipped'] = skipped
    metadata['excel_errors'] = errors

    with open(metadata_file, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, indent=2)

    # Print summary
    print(f"\n--- Excel Extraction Complete ---")
    print(f"Processed: {processed} files")
    print(f"Skipped (already extracted): {skipped} files")
    print(f"Errors: {errors} files")
    print(f"Total sheets extracted: {total_sheets}")
    print(f"Index saved to: {index_folder}")


def main():
    parser = argparse.ArgumentParser(description='Extract text from Excel files for searching')
    parser.add_argument('--source', help='Source folder containing Excel files (overrides config.json)')
    parser.add_argument('--index', help='Index folder for extracted text (overrides config.json)')
    parser.add_argument('--reindex', action='store_true', help='Force re-extract all files')
    args = parser.parse_args()

    # Load config
    config = load_config()

    # Get settings (command line args override config)
    source_folder = args.source or config.get('excel_source_folder')
    index_folder = args.index or config.get('index_folder', './index')
    extensions = config.get('excel_extensions', EXCEL_EXTENSIONS)

    if not source_folder:
        print("Error: No Excel source folder specified.")
        print("Set 'excel_source_folder' in config.json or use --source argument.")
        sys.exit(1)

    print(f"Source: {source_folder}")
    print(f"Index:  {index_folder}")
    print(f"Extensions: {extensions}")
    print(f"Reindex: {args.reindex}")
    print()

    extract_all(source_folder, index_folder, reindex=args.reindex, extensions=extensions)


if __name__ == '__main__':
    main()
