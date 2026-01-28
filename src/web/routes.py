"""Search routes and logic."""

import os
import sys
import json
import re
import threading
from markupsafe import Markup
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, current_app, send_file, abort
from .auth import login_required, check_password, authenticate_user, logout_user
from .blob_storage import is_blob_storage_enabled, generate_pdf_sas_url, check_blob_exists, download_index_from_blob, is_index_download_complete, is_index_downloading, get_blob_service_client, PDF_CONTAINER
from .projects import get_project, get_all_projects
from . import get_index_folder
from .llm_query import (
    parse_query_with_llm, QueryPlan, LLMError, LLMTimeoutError, LLMValidationError,
    is_smart_search_enabled, is_keyword_search_enabled
)
from .query_logger import log_search
from .admin import (
    admin_required, is_admin, get_current_user_email, get_current_user_name,
    list_app_users, add_user_access, remove_user_access, GraphAPIError
)

main = Blueprint('main', __name__)

# Per-project index storage (loaded on first request per project)
_indexes = {}    # {project_id: [doc, doc, ...]}
_metadatas = {}  # {project_id: {total_docs: N, ...}}
_preload_states = {}  # {project_id: {'in_progress': bool, 'complete': bool}}
_preload_lock = threading.Lock()


def load_project_index(project_id):
    """Load all extracted text into memory for a specific project."""
    global _indexes, _metadatas

    if project_id in _indexes:
        return _indexes[project_id], _metadatas[project_id]

    project = get_project(project_id)
    if project is None:
        abort(404, 'Project not found')

    config_folder = project.get('index_folder', f'./index/{project_id}')
    index_folder = get_index_folder(project_id, config_folder)
    texts_folder = os.path.join(index_folder, 'texts')
    metadata_file = os.path.join(index_folder, 'metadata.json')

    # If blob storage is enabled and index is still downloading, show loading
    if is_blob_storage_enabled() and is_index_downloading(project_id) and not is_index_download_complete(project_id):
        return [], {'total_docs': 0, 'total_pages': 0, 'loading': True, 'loading_message': 'Downloading index files...'}

    # If index is being pre-loaded in background, show loading
    if is_index_preloading(project_id):
        return [], {'total_docs': 0, 'total_pages': 0, 'loading': True, 'loading_message': 'Loading documents into memory...'}

    index = []
    metadata = {'total_docs': 0, 'total_pages': 0}

    # Load metadata if exists
    if os.path.exists(metadata_file):
        with open(metadata_file, 'r', encoding='utf-8') as f:
            metadata = json.load(f)

    # Load all text files
    if os.path.exists(texts_folder):
        for filename in os.listdir(texts_folder):
            if filename.endswith('.json'):
                filepath = os.path.join(texts_folder, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        doc = json.load(f)
                        # Default file_type for legacy PDF-only index files
                        if 'file_type' not in doc:
                            doc['file_type'] = 'pdf'
                        index.append(doc)
                except Exception as e:
                    print(f"Error loading {filepath}: {e}")

    metadata['total_docs'] = len(index)
    metadata['total_pages'] = sum(len(doc.get('pages', [])) for doc in index)

    _indexes[project_id] = index
    _metadatas[project_id] = metadata

    return index, metadata


def is_index_preloading(project_id):
    """Check if index is currently being pre-loaded."""
    with _preload_lock:
        state = _preload_states.get(project_id, {})
        return state.get('in_progress', False)


def is_index_preload_complete(project_id):
    """Check if index pre-loading has completed."""
    with _preload_lock:
        state = _preload_states.get(project_id, {})
        return state.get('complete', False)


def _preload_single_index(project_id, index_folder):
    """Pre-load a single project's index (runs in background thread)."""
    global _indexes, _metadatas

    with _preload_lock:
        _preload_states[project_id] = {'in_progress': True, 'complete': False}

    try:
        texts_folder = os.path.join(index_folder, 'texts')
        metadata_file = os.path.join(index_folder, 'metadata.json')

        index = []
        metadata = {'total_docs': 0, 'total_pages': 0}

        # Load metadata if exists
        if os.path.exists(metadata_file):
            with open(metadata_file, 'r', encoding='utf-8') as f:
                metadata = json.load(f)

        # Load all text files
        if os.path.exists(texts_folder):
            files = [f for f in os.listdir(texts_folder) if f.endswith('.json')]
            print(f"Pre-loading {len(files)} documents for project '{project_id}'...", file=sys.stderr, flush=True)

            for i, filename in enumerate(files):
                filepath = os.path.join(texts_folder, filename)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        doc = json.load(f)
                        if 'file_type' not in doc:
                            doc['file_type'] = 'pdf'
                        index.append(doc)
                except Exception as e:
                    print(f"Error loading {filepath}: {e}", file=sys.stderr, flush=True)

                # Progress logging for large indices
                if (i + 1) % 5000 == 0:
                    print(f"  Pre-loaded {i + 1}/{len(files)} documents for '{project_id}'...", file=sys.stderr, flush=True)

        metadata['total_docs'] = len(index)
        metadata['total_pages'] = sum(len(doc.get('pages', [])) for doc in index)

        _indexes[project_id] = index
        _metadatas[project_id] = metadata

        print(f"Pre-load complete for project '{project_id}': {len(index)} documents", file=sys.stderr, flush=True)

    except Exception as e:
        print(f"Error pre-loading index for '{project_id}': {e}", file=sys.stderr, flush=True)
    finally:
        with _preload_lock:
            _preload_states[project_id] = {'in_progress': False, 'complete': True}


def start_index_preload(project_id, index_folder):
    """Start pre-loading an index in a background thread."""
    # Skip if already loaded or loading
    if project_id in _indexes:
        print(f"Index for '{project_id}' already in memory, skipping pre-load", file=sys.stderr, flush=True)
        return

    with _preload_lock:
        if _preload_states.get(project_id, {}).get('in_progress', False):
            print(f"Index for '{project_id}' already pre-loading, skipping", file=sys.stderr, flush=True)
            return

    thread = threading.Thread(
        target=_preload_single_index,
        args=(project_id, index_folder),
        daemon=True
    )
    thread.start()


def parse_search_terms(query):
    """Parse search query into individual terms.

    Supports:
    - Multiple words: 'Ford Confidential' -> ['ford', 'confidential'] (AND search)
    - Quoted phrases: '"exact phrase"' -> ['exact phrase'] (exact match)
    """
    terms = []
    query = query.strip()

    # Extract quoted phrases first
    quoted_pattern = r'"([^"]+)"'
    quoted_matches = re.findall(quoted_pattern, query)
    terms.extend([m.lower() for m in quoted_matches])

    # Remove quoted phrases from query
    remaining = re.sub(quoted_pattern, '', query).strip()

    # Split remaining by whitespace
    if remaining:
        words = remaining.split()
        terms.extend([w.lower() for w in words if w])

    return terms


def search_index(query, project_id, page=1, per_page=20, file_type_filter=None):
    """Search the index for matching documents within a project.

    Smart search: Multiple words are treated as AND search.
    Use quotes for exact phrase matching: "Ford Confidential"
    file_type_filter: None (all), 'pdf', or 'excel'
    """
    index, metadata = load_project_index(project_id)

    if not query or not query.strip():
        return {
            'query': '',
            'total_matches': 0,
            'documents': 0,
            'results': [],
            'page': page,
            'total_pages': 0,
            'has_more': False
        }

    search_terms = parse_search_terms(query)
    if not search_terms:
        return {
            'query': query,
            'total_matches': 0,
            'documents': 0,
            'results': [],
            'page': page,
            'total_pages': 0,
            'has_more': False
        }

    results = []
    seen_docs = set()

    for doc in index:
        # Apply file type filter
        doc_type = doc.get('file_type', 'pdf')
        if file_type_filter and doc_type != file_type_filter:
            continue
        filename = doc.get('filename', '')
        filepath = doc.get('path', '')

        for page_info in doc.get('pages', []):
            page_num = page_info.get('page_num', 0)
            text = page_info.get('text', '')
            text_lower = text.lower()

            # Check if ALL search terms are present (AND search)
            if all(term in text_lower for term in search_terms):
                # Extract context around the first matching term
                context = extract_context(text, search_terms[0])

                # Count total matches across all terms
                match_count = sum(text_lower.count(term) for term in search_terms)

                results.append({
                    'filename': filename,
                    'filepath': filepath,
                    'page': page_num,
                    'sheet_name': page_info.get('sheet_name', ''),
                    'file_type': doc.get('file_type', 'pdf'),
                    'context': context,
                    'match_count': match_count
                })
                seen_docs.add(filename)

    # Sort by match count (most matches first)
    results.sort(key=lambda x: x['match_count'], reverse=True)

    # Pagination
    total_matches = len(results)
    total_pages = (total_matches + per_page - 1) // per_page  # Ceiling division
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_results = results[start_idx:end_idx]

    return {
        'query': query,
        'total_matches': total_matches,
        'documents': len(seen_docs),
        'results': paginated_results,
        'page': page,
        'total_pages': total_pages,
        'has_more': end_idx < total_matches
    }


def extract_context(text, query, context_chars=100):
    """Extract a snippet of text around the first match."""
    text_lower = text.lower()
    pos = text_lower.find(query)

    if pos == -1:
        return text[:200] + '...' if len(text) > 200 else text

    # Get context before and after
    start = max(0, pos - context_chars)
    end = min(len(text), pos + len(query) + context_chars)

    snippet = text[start:end]

    # Add ellipsis if truncated
    if start > 0:
        snippet = '...' + snippet
    if end < len(text):
        snippet = snippet + '...'

    return snippet


def highlight_matches(text, search_terms):
    """Highlight all search terms in text (case-insensitive)."""
    if not search_terms:
        return text

    for term in search_terms:
        pattern = re.compile(re.escape(term), re.IGNORECASE)
        text = pattern.sub(lambda m: f'<mark>{m.group()}</mark>', text)

    return text


def smart_search_index(query_plan, project_id, page=1, per_page=20, file_type_filter=None):
    """Search the index using a QueryPlan from LLM parsing.

    Uses OR logic for optional terms and weighted scoring.
    """
    index, metadata = load_project_index(project_id)

    if not query_plan.required_terms:
        return {
            'query': '',
            'total_matches': 0,
            'documents': 0,
            'results': [],
            'page': page,
            'total_pages': 0,
            'has_more': False
        }

    results = []
    seen_docs = set()

    # Combine all search terms for matching
    required_terms = query_plan.required_terms
    optional_terms = query_plan.optional_terms
    person_names = query_plan.person_names
    locations = query_plan.locations

    for doc in index:
        # Apply file type filter
        doc_type = doc.get('file_type', 'pdf')
        if file_type_filter and doc_type != file_type_filter:
            continue
        filename = doc.get('filename', '')
        filepath = doc.get('path', '')

        for page_info in doc.get('pages', []):
            page_num = page_info.get('page_num', 0)
            text = page_info.get('text', '')
            text_lower = text.lower()

            # Check if ALL required terms are present
            if not all(term in text_lower for term in required_terms):
                continue

            # Calculate relevance score
            score = 0.0

            # Required terms (base score)
            for term in required_terms:
                count = text_lower.count(term)
                score += min(count * 10, 30)  # Cap per term

            # Optional/synonym terms (bonus)
            for term in optional_terms:
                if term in text_lower:
                    score += 5

            # Person name matches (high value)
            for name in person_names:
                if name in text_lower:
                    score += 15

            # Location matches
            for loc in locations:
                if loc in text_lower:
                    score += 10

            # Normalize to 0-100
            score = min(score, 100)

            # Extract context around the first required term
            context = extract_context(text, required_terms[0])

            results.append({
                'filename': filename,
                'filepath': filepath,
                'page': page_num,
                'sheet_name': page_info.get('sheet_name', ''),
                'file_type': doc.get('file_type', 'pdf'),
                'context': context,
                'match_count': int(score),  # Use score as match count for sorting
                'relevance_score': score
            })
            seen_docs.add(filename)

    # Sort by relevance score (highest first)
    results.sort(key=lambda x: x['relevance_score'], reverse=True)

    # Pagination
    total_matches = len(results)
    total_pages = (total_matches + per_page - 1) // per_page
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_results = results[start_idx:end_idx]

    return {
        'query': ' '.join(required_terms),
        'total_matches': total_matches,
        'documents': len(seen_docs),
        'results': paginated_results,
        'page': page,
        'total_pages': total_pages,
        'has_more': end_idx < total_matches
    }


def _get_project_or_404(project_id):
    """Validate project_id and return project config, or abort 404."""
    project = get_project(project_id)
    if project is None:
        abort(404, 'Project not found')
    return project


# ─── Auth routes ───

@main.route('/login', methods=['GET', 'POST'])
def login():
    """Login page."""
    if request.method == 'POST':
        password = request.form.get('password', '')
        if check_password(password):
            authenticate_user()
            next_url = request.args.get('next', url_for('main.project_picker'))
            return redirect(next_url)
        flash('Invalid password', 'error')

    return render_template('login.html')


@main.route('/logout')
def logout():
    """Logout and redirect to login."""
    logout_user()
    return redirect(url_for('main.login'))


# ─── Project picker ───

@main.route('/')
@login_required
def project_picker():
    """Show project selection page, or redirect if only one project."""
    projects = get_all_projects()
    if len(projects) == 1:
        return redirect(url_for('main.project_search', project_id=projects[0]['id']))
    return render_template('projects.html', projects=projects)


@main.route('/api/projects')
@login_required
def api_projects():
    """List available projects."""
    return jsonify(get_all_projects())


# ─── Project-scoped routes ───

@main.route('/p/<project_id>/')
@login_required
def project_search(project_id):
    """Main search page for a specific project."""
    project = _get_project_or_404(project_id)
    _, metadata = load_project_index(project_id)
    return render_template('search.html', metadata=metadata, project=project)


@main.route('/p/<project_id>/api/search')
@login_required
def project_api_search(project_id):
    """Keyword search API endpoint scoped to a project."""
    _get_project_or_404(project_id)
    query = request.args.get('q', '')
    page = request.args.get('page', 1, type=int)
    file_type = request.args.get('type', '')  # '', 'pdf', or 'excel'

    results = search_index(query, project_id=project_id, page=page, file_type_filter=file_type or None)

    # Highlight matches in context (using parsed search terms)
    search_terms = parse_search_terms(query)
    for result in results['results']:
        result['context'] = highlight_matches(result['context'], search_terms)

    # Log the search
    log_search(
        project_id=project_id,
        query_text=query,
        search_mode='keyword',
        result_count=results['total_matches']
    )

    return jsonify(results)


@main.route('/p/<project_id>/api/smart-search', methods=['POST'])
@login_required
def project_api_smart_search(project_id):
    """Smart search API endpoint using LLM query parsing."""
    import sys
    import time

    project = _get_project_or_404(project_id)
    data = request.get_json() or {}
    query = data.get('query', '').strip()
    page = data.get('page', 1)
    file_type = data.get('type', '')

    print(f"[SMART-SEARCH] Request received: project={project_id}, query='{query}'", file=sys.stderr)

    if not query:
        return jsonify({
            'error': 'Please enter a search query',
            'query': '',
            'total_matches': 0,
            'documents': 0,
            'results': [],
            'page': 1,
            'total_pages': 0,
            'has_more': False
        })

    # Check if smart search is enabled
    if not is_smart_search_enabled():
        return jsonify({
            'error': 'Smart search is not enabled',
            'fallback_available': is_keyword_search_enabled()
        }), 503

    start_time = time.time()
    llm_latency_ms = None
    error_message = None
    query_plan = None
    cache_hit = False

    try:
        # Parse the query with LLM
        print(f"[SMART-SEARCH] Calling LLM...", file=sys.stderr)
        llm_start = time.time()
        query_plan = parse_query_with_llm(
            query,
            project_id=project_id,
            project_description=project.get('description', '')
        )
        llm_latency_ms = int((time.time() - llm_start) * 1000)
        print(f"[SMART-SEARCH] LLM response in {llm_latency_ms}ms: {query_plan.required_terms}", file=sys.stderr)

        # Check if it was a cache hit (latency < 50ms suggests cache)
        cache_hit = llm_latency_ms < 50

        # Perform the search
        results = smart_search_index(
            query_plan,
            project_id=project_id,
            page=page,
            file_type_filter=file_type or None
        )

        # Highlight matches in context
        all_terms = (
            query_plan.required_terms +
            query_plan.optional_terms +
            query_plan.person_names +
            query_plan.locations
        )
        for result in results['results']:
            result['context'] = highlight_matches(result['context'], all_terms)

        # Add smart search metadata to response
        results['smart_search'] = True
        results['interpretation'] = query_plan.interpretation
        results['query_plan'] = {
            'required_terms': query_plan.required_terms,
            'optional_terms': query_plan.optional_terms,
            'person_names': query_plan.person_names,
            'locations': query_plan.locations,
            'confidence': query_plan.confidence
        }
        results['llm_latency_ms'] = llm_latency_ms
        results['cache_hit'] = cache_hit

        # Add confidence warning if low
        if query_plan.confidence < 0.7:
            results['warning'] = 'Results may not be precise - query was ambiguous'

        # Log the search
        log_search(
            project_id=project_id,
            query_text=query,
            search_mode='smart',
            query_plan=query_plan.to_dict(),
            interpretation=query_plan.interpretation,
            result_count=results['total_matches'],
            llm_latency_ms=llm_latency_ms,
            cache_hit=cache_hit
        )

        return jsonify(results)

    except LLMTimeoutError as e:
        error_message = str(e)
        print(f"[SMART-SEARCH] LLM TIMEOUT: {error_message}", file=sys.stderr)
        log_search(
            project_id=project_id,
            query_text=query,
            search_mode='smart',
            error_message=error_message,
            llm_latency_ms=llm_latency_ms
        )
        return jsonify({
            'error': 'Search is taking longer than expected. Please try again.',
            'error_type': 'timeout',
            'fallback_available': is_keyword_search_enabled()
        }), 504

    except LLMValidationError as e:
        error_message = str(e)
        print(f"[SMART-SEARCH] LLM VALIDATION ERROR: {error_message}", file=sys.stderr)
        log_search(
            project_id=project_id,
            query_text=query,
            search_mode='smart',
            error_message=error_message,
            llm_latency_ms=llm_latency_ms
        )
        return jsonify({
            'error': f"Couldn't understand query: {e}. Try rephrasing.",
            'error_type': 'validation',
            'fallback_available': is_keyword_search_enabled()
        }), 400

    except LLMError as e:
        error_message = str(e)
        print(f"[SMART-SEARCH] LLM ERROR: {error_message}", file=sys.stderr)
        log_search(
            project_id=project_id,
            query_text=query,
            search_mode='smart',
            error_message=error_message,
            llm_latency_ms=llm_latency_ms
        )
        return jsonify({
            'error': 'Search service temporarily unavailable. Please try again.',
            'error_type': 'service_error',
            'fallback_available': is_keyword_search_enabled()
        }), 503

    except Exception as e:
        import traceback
        error_message = str(e)
        print(f"[SMART-SEARCH] UNEXPECTED ERROR: {error_message}", file=sys.stderr)
        print(f"[SMART-SEARCH] Traceback: {traceback.format_exc()}", file=sys.stderr)
        current_app.logger.error(f"Smart search error: {e}")
        log_search(
            project_id=project_id,
            query_text=query,
            search_mode='smart',
            error_message=error_message,
            llm_latency_ms=llm_latency_ms
        )
        return jsonify({
            'error': 'An unexpected error occurred. Please try again.',
            'error_type': 'unknown',
            'fallback_available': is_keyword_search_enabled()
        }), 500


@main.route('/p/<project_id>/api/search-config')
@login_required
def project_search_config(project_id):
    """Get search configuration for the frontend."""
    _get_project_or_404(project_id)
    return jsonify({
        'smart_search_enabled': is_smart_search_enabled(),
        'keyword_search_enabled': is_keyword_search_enabled(),
        'default_mode': 'smart' if is_smart_search_enabled() else 'keyword'
    })


@main.route('/p/<project_id>/api/stats')
@login_required
def project_api_stats(project_id):
    """Get index statistics for a project."""
    _get_project_or_404(project_id)
    _, metadata = load_project_index(project_id)
    return jsonify(metadata)


@main.route('/health')
def health():
    """Health check endpoint for container orchestration."""
    projects = get_all_projects()
    project_statuses = {}
    for p in projects:
        pid = p['id']
        project_statuses[pid] = {
            'index_downloading': is_index_downloading(pid),
            'index_ready': is_index_download_complete(pid)
        }
    return jsonify({
        'status': 'ok',
        'projects': project_statuses
    })


@main.route('/p/<project_id>/pdf/<path:filepath>')
@login_required
def project_serve_pdf(project_id, filepath):
    """Serve a PDF file - from blob storage or local source folder."""
    project = _get_project_or_404(project_id)

    # If blob storage is configured, redirect to SAS URL
    if is_blob_storage_enabled():
        current_app.logger.info(f"PDF request (blob): project={project_id}, filepath={filepath!r}")

        if not check_blob_exists(filepath):
            current_app.logger.error(f"File not found in blob storage: {filepath}")
            abort(404, 'File not found in blob storage')

        sas_url = generate_pdf_sas_url(filepath, expiry_hours=1)
        return redirect(sas_url)

    # Fallback to local file serving
    source_folder = project.get('source_folder', '')
    current_app.logger.info(f"PDF request (local): project={project_id}, filepath={filepath!r}, source_folder={source_folder!r}")

    if not source_folder:
        current_app.logger.error(f"source_folder not configured for project '{project_id}'")
        abort(404, 'Source folder not configured')

    # Construct full path and ensure it's within source folder (security)
    full_path = os.path.normpath(os.path.join(source_folder, filepath))
    current_app.logger.info(f"Full path: {full_path!r}, exists={os.path.exists(full_path)}")

    if not full_path.startswith(os.path.normpath(source_folder)):
        abort(403, 'Access denied')

    if not os.path.exists(full_path):
        current_app.logger.error(f"File not found: {full_path}")
        abort(404, 'File not found')

    return send_file(full_path, mimetype='application/pdf')


@main.route('/p/<project_id>/doc-view/<path:filepath>')
@login_required
def project_doc_view(project_id, filepath):
    """View a Word document or email as rendered text."""
    project = _get_project_or_404(project_id)
    query = request.args.get('q', '')

    # Find the corresponding index file to get the extracted text
    index_folder = project.get('index_folder', f'./index/{project_id}')
    filename = os.path.basename(filepath)

    # Try to find the text file in the index
    text_file = os.path.join(index_folder, 'texts', f'{filename}.json')

    if not os.path.exists(text_file):
        # If local file doesn't exist and blob storage is enabled, try to get from loaded index
        index, _ = load_project_index(project_id)
        doc_data = None
        for doc in index:
            if doc.get('path', '').replace('\\', '/') == filepath.replace('\\', '/'):
                doc_data = doc
                break

        if not doc_data:
            abort(404, 'Document not found in index')

        # Get content from the loaded index
        pages = doc_data.get('pages', [])
        content = '\n\n'.join(p.get('text', '') for p in pages)
        file_type = doc_data.get('file_type', 'word')
    else:
        # Load from local file
        with open(text_file, 'r', encoding='utf-8') as f:
            doc_data = json.load(f)
        pages = doc_data.get('pages', [])
        content = '\n\n'.join(p.get('text', '') for p in pages)
        file_type = doc_data.get('file_type', 'word')

    # Highlight search terms if provided
    if query:
        search_terms = parse_search_terms(query)
        for term in search_terms:
            import re
            pattern = re.compile(re.escape(term), re.IGNORECASE)
            content = pattern.sub(lambda m: f'<mark>{m.group()}</mark>', content)

    # Escape HTML except for our mark tags
    content = content.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    content = content.replace('&lt;mark&gt;', '<mark>').replace('&lt;/mark&gt;', '</mark>')

    # Extract email metadata if available
    email_meta = None
    if file_type == 'email':
        email_meta = {
            'subject': doc_data.get('subject', ''),
            'sender': doc_data.get('sender', ''),
            'to': doc_data.get('to', ''),
            'date': doc_data.get('date', '')
        }

    return render_template('doc_viewer.html',
        filename=filename,
        filepath=filepath,
        content=content,
        file_type=file_type,
        email_meta=email_meta,
        project_id=project_id,
        query=query
    )


def _resolve_excel_path(filepath, project):
    """Resolve an Excel filepath to a local file path, with security checks.

    If blob storage is enabled, downloads the file to a temp location first.
    Returns the local file path (either from project's excel_source_folder or temp download).
    """
    # If blob storage is enabled, download from blob to temp file
    if is_blob_storage_enabled():
        if not check_blob_exists(filepath):
            current_app.logger.error(f"Excel file not found in blob storage: {filepath}")
            abort(404, 'File not found in blob storage')

        import tempfile
        blob_service_client = get_blob_service_client()
        blob_client = blob_service_client.get_blob_client(PDF_CONTAINER, filepath)

        # Download to temp file preserving the extension
        ext = os.path.splitext(filepath)[1]
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        try:
            tmp.write(blob_client.download_blob().readall())
            tmp.close()
            return tmp.name
        except Exception as e:
            tmp.close()
            os.unlink(tmp.name)
            current_app.logger.error(f"Error downloading Excel from blob: {e}")
            abort(500, 'Error downloading file from storage')

    # Local file serving
    excel_source = project.get('excel_source_folder', '')

    if not excel_source:
        current_app.logger.error(f"excel_source_folder not configured for project '{project['id']}'")
        abort(404, 'Excel source folder not configured')

    full_path = os.path.normpath(os.path.join(excel_source, filepath))

    # Security: ensure path is within source folder
    if not full_path.startswith(os.path.normpath(excel_source)):
        abort(403, 'Access denied')

    if not os.path.exists(full_path):
        current_app.logger.error(f"Excel file not found: {full_path}")
        abort(404, 'File not found')

    return full_path


def _open_excel_workbook(full_path):
    """Open an Excel file and return (sheets_data, sheet_names).

    sheets_data is a dict: {sheet_name: list of rows}, where each row is a list of cell values.
    """
    ext = os.path.splitext(full_path)[1].lower()
    sheets_data = {}
    sheet_names = []

    if ext in ('.xlsx', '.xlsm'):
        import openpyxl
        from openpyxl.chartsheet import Chartsheet
        wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
        try:
            sheet_names = wb.sheetnames
            for sname in sheet_names:
                ws = wb[sname]
                if isinstance(ws, Chartsheet):
                    continue
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([c for c in row])
                sheets_data[sname] = rows
        finally:
            wb.close()

    elif ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(full_path)
        sheet_names = wb.sheet_names()
        for sname in sheet_names:
            ws = wb.sheet_by_name(sname)
            rows = []
            for r in range(ws.nrows):
                row = [ws.cell_value(r, c) for c in range(ws.ncols)]
                rows.append(row)
            sheets_data[sname] = rows

    elif ext == '.xlsb':
        from pyxlsb import open_workbook
        with open_workbook(full_path) as wb:
            sheet_names = wb.sheets
            for sname in sheet_names:
                with wb.get_sheet(sname) as ws:
                    rows = []
                    for row in ws.rows():
                        rows.append([c.v for c in row])
                    sheets_data[sname] = rows
    else:
        abort(400, f'Unsupported file type: {ext}')

    return sheets_data, sheet_names


def _render_sheet_html(rows, search_terms, full_mode=False, context_rows=5):
    """Render sheet rows as an HTML table with highlighted matches.

    In contextual mode (full_mode=False): shows header + rows around matches.
    In full mode: shows all rows.
    Returns (html_string, match_count).
    """
    if not rows:
        return '<p class="no-data">This sheet is empty.</p>', 0

    # Detect header row
    first_row = rows[0]
    has_header = False
    non_empty = [c for c in first_row if c is not None and str(c).strip()]
    if non_empty:
        string_count = sum(1 for c in non_empty if isinstance(c, str))
        has_header = string_count / len(non_empty) >= 0.6 if non_empty else False

    header_row = rows[0] if has_header else None
    data_rows = rows[1:] if has_header else rows

    if full_mode:
        # Show all rows
        visible_indices = set(range(len(data_rows)))
    else:
        # Find rows containing any search term
        match_indices = []
        for i, row in enumerate(data_rows):
            row_text = ' '.join(str(c) for c in row if c is not None).lower()
            if any(term in row_text for term in search_terms):
                match_indices.append(i)

        # Build visible set: context_rows before and after each match
        visible_indices = set()
        for idx in match_indices:
            for offset in range(-context_rows, context_rows + 1):
                adj = idx + offset
                if 0 <= adj < len(data_rows):
                    visible_indices.add(adj)

        # If no matches found in data rows, show first 20 rows
        if not visible_indices:
            visible_indices = set(range(min(20, len(data_rows))))

    # Build HTML table
    parts = ['<table class="excel-table">']

    # Header row
    if header_row:
        parts.append('<thead><tr>')
        for cell in header_row:
            cell_str = str(cell) if cell is not None else ''
            cell_html = _escape_html(cell_str)
            if search_terms:
                cell_html = highlight_matches(cell_html, search_terms)
            parts.append(f'<th>{cell_html}</th>')
        parts.append('</tr></thead>')

    # Data rows
    parts.append('<tbody>')
    match_count = 0
    prev_idx = -2  # Track gaps for dividers
    sorted_indices = sorted(visible_indices)

    for idx in sorted_indices:
        # Insert divider if there's a gap
        if idx > prev_idx + 1 and prev_idx >= 0:
            col_count = len(header_row) if header_row else len(data_rows[idx]) if data_rows[idx] else 1
            parts.append(f'<tr class="match-divider"><td colspan="{col_count}">...</td></tr>')

        row = data_rows[idx]
        row_text = ' '.join(str(c) for c in row if c is not None).lower()
        is_match = any(term in row_text for term in search_terms)
        row_class = ' class="match-row"' if is_match else ''

        if is_match:
            match_count += 1

        parts.append(f'<tr{row_class}>')
        for cell in row:
            cell_str = str(cell) if cell is not None else ''
            if cell_str == 'None':
                cell_str = ''
            cell_html = _escape_html(cell_str)
            if search_terms and cell_str:
                cell_html = highlight_matches(cell_html, search_terms)
            parts.append(f'<td>{cell_html}</td>')
        parts.append('</tr>')
        prev_idx = idx

    parts.append('</tbody></table>')

    total_rows = len(data_rows)
    shown_rows = len(sorted_indices)
    if not full_mode and shown_rows < total_rows:
        parts.append(f'<p class="truncation-notice">Showing {shown_rows} of {total_rows} rows (rows near matches). <a href="?full=1&sheet={{}}&q={{}}" class="view-full-link">View entire sheet</a></p>')

    return '\n'.join(parts), match_count


def _escape_html(text):
    """Escape HTML special characters."""
    return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


@main.route('/p/<project_id>/excel-view/<path:filepath>')
@login_required
def project_excel_view(project_id, filepath):
    """View an Excel sheet as an HTML table with search term highlighting."""
    project = _get_project_or_404(project_id)
    sheet_name = request.args.get('sheet', '')
    query = request.args.get('q', '')
    full_mode = request.args.get('full', '0') == '1'

    full_path = _resolve_excel_path(filepath, project)
    is_temp = is_blob_storage_enabled()

    try:
        sheets_data, sheet_names = _open_excel_workbook(full_path)
    finally:
        # Clean up temp file from blob download
        if is_temp and os.path.exists(full_path):
            os.unlink(full_path)

    # Default to first sheet if not specified or invalid
    if not sheet_name or sheet_name not in sheets_data:
        sheet_name = sheet_names[0] if sheet_names else ''

    search_terms = parse_search_terms(query) if query else []

    # Render the requested sheet
    rows = sheets_data.get(sheet_name, [])
    table_html, match_count = _render_sheet_html(rows, search_terms, full_mode=full_mode)

    # Fix the view-full-link placeholders
    from urllib.parse import quote
    table_html = table_html.replace(
        'sheet={}&q={}',
        f'sheet={quote(sheet_name)}&q={quote(query)}'
    )

    return render_template('excel_view.html',
        filename=os.path.basename(filepath),
        filepath=filepath,
        sheet_names=sheet_names,
        active_sheet=sheet_name,
        query=query,
        full_mode=full_mode,
        match_count=match_count,
        table_html=Markup(table_html),
        total_rows=len(rows),
        project=project
    )


EXCEL_MIME_TYPES = {
    '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    '.xls': 'application/vnd.ms-excel',
    '.xlsm': 'application/vnd.ms-excel.sheet.macroEnabled.12',
    '.xlsb': 'application/vnd.ms-excel.sheet.binary.macroEnabled.12',
}


@main.route('/p/<project_id>/file/<path:filepath>')
@login_required
def project_serve_file(project_id, filepath):
    """Serve an Excel file for download."""
    project = _get_project_or_404(project_id)
    ext = os.path.splitext(filepath)[1].lower()

    # For blob storage, generate SAS URL
    if is_blob_storage_enabled():
        if not check_blob_exists(filepath):
            abort(404, 'File not found in blob storage')
        sas_url = generate_pdf_sas_url(filepath, expiry_hours=1)
        return redirect(sas_url)

    # Local file serving
    full_path = _resolve_excel_path(filepath, project)

    mimetype = EXCEL_MIME_TYPES.get(ext, 'application/octet-stream')
    return send_file(full_path, mimetype=mimetype,
                     as_attachment=True,
                     download_name=os.path.basename(filepath))


# ─── Backward compatibility routes ───
# These redirect old URLs to the new project-scoped URLs when only one project exists.

def _get_single_project_id():
    """Get the project ID if there's only one project, otherwise abort 404."""
    projects = get_all_projects()
    if len(projects) == 1:
        return projects[0]['id']
    abort(404)


@main.route('/api/search')
@login_required
def legacy_api_search():
    """Legacy search API — redirects to single project."""
    project_id = _get_single_project_id()
    return redirect(url_for('main.project_api_search', project_id=project_id, **request.args))


@main.route('/api/stats')
@login_required
def legacy_api_stats():
    """Legacy stats API — redirects to single project."""
    project_id = _get_single_project_id()
    return redirect(url_for('main.project_api_stats', project_id=project_id))


@main.route('/pdf/<path:filepath>')
@login_required
def legacy_serve_pdf(filepath):
    """Legacy PDF serving — redirects to single project."""
    project_id = _get_single_project_id()
    return redirect(url_for('main.project_serve_pdf', project_id=project_id, filepath=filepath))


@main.route('/excel-view/<path:filepath>')
@login_required
def legacy_excel_view(filepath):
    """Legacy Excel view — redirects to single project."""
    project_id = _get_single_project_id()
    return redirect(url_for('main.project_excel_view', project_id=project_id, filepath=filepath, **request.args))


@main.route('/file/<path:filepath>')
@login_required
def legacy_serve_file(filepath):
    """Legacy file download — redirects to single project."""
    project_id = _get_single_project_id()
    return redirect(url_for('main.project_serve_file', project_id=project_id, filepath=filepath))


# ─── Admin Routes ───────────────────────────────────────────────────────────

@main.route('/admin')
@login_required
@admin_required
def admin_page():
    """Admin page for user management."""
    return render_template('admin.html')


@main.route('/admin/api/users', methods=['GET'])
@login_required
@admin_required
def admin_list_users():
    """API endpoint to list users with access."""
    try:
        users = list_app_users()
        return jsonify({'users': users})
    except GraphAPIError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        current_app.logger.error(f"Error listing users: {e}")
        return jsonify({'error': 'Failed to load users'}), 500


@main.route('/admin/api/users', methods=['POST'])
@login_required
@admin_required
def admin_add_user():
    """API endpoint to add a user."""
    data = request.get_json() or {}
    email = data.get('email', '').strip().lower()

    if not email:
        return jsonify({'error': 'Email is required'}), 400

    # Validate Thompsons domain
    if not email.endswith('@thompsons-scotland.co.uk') and not email.endswith('@thompsons.co.uk'):
        return jsonify({'error': 'Only Thompsons email addresses are allowed'}), 400

    try:
        result = add_user_access(email)
        return jsonify(result)
    except GraphAPIError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        current_app.logger.error(f"Error adding user {email}: {e}")
        return jsonify({'error': 'Failed to add user'}), 500


@main.route('/admin/api/users/<assignment_id>', methods=['DELETE'])
@login_required
@admin_required
def admin_remove_user(assignment_id):
    """API endpoint to remove a user's access."""
    try:
        remove_user_access(assignment_id)
        return jsonify({'status': 'removed'})
    except GraphAPIError as e:
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        current_app.logger.error(f"Error removing user {assignment_id}: {e}")
        return jsonify({'error': 'Failed to remove user'}), 500
